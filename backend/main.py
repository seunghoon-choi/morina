"""
ByeTax API - FastAPI 메인 앱
PDF 업로드 → 파싱 → SQLite 저장 → 조회
"""
import io
import os
import shutil
import uuid
from datetime import datetime, timedelta
from urllib.parse import quote

from fastapi import Depends, FastAPI, File, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from auth import (
    KAKAO_CLIENT_ID,
    create_access_token,
    get_current_user,
    get_kakao_login_url,
    get_kakao_profile,
    get_kakao_token,
    get_user_by_id,
    upsert_user,
)
from db import get_conn, init_db
from pdf_parser import parse_tax_pdf
from tax_calculator import calculate_tax, generate_ai_analysis

FRONTEND_DIR = os.path.join(os.path.dirname(__file__), "..", "frontend")
UPLOAD_DIR   = os.path.join(os.path.dirname(__file__), "..", "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)


def _asset_version() -> str:
    """정적 파일 mtime 기반 캐시 버스팅 버전"""
    try:
        css = os.path.getmtime(os.path.join(FRONTEND_DIR, "css", "style.css"))
        js  = os.path.getmtime(os.path.join(FRONTEND_DIR, "js", "app.js"))
        return str(int(max(css, js)))
    except OSError:
        return "1"

app = FastAPI(title="ByeTax API", version="0.2.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def startup():
    init_db()


# ── 프론트엔드 정적 파일 서빙 ─────────────────────────────────────────────────
app.mount("/static", StaticFiles(directory=FRONTEND_DIR), name="static")

@app.get("/", include_in_schema=False)
def serve_index():
    html_path = os.path.join(FRONTEND_DIR, "index.html")
    with open(html_path, "r", encoding="utf-8") as f:
        html = f.read()
    v = _asset_version()
    html = html.replace("style.css\"", f"style.css?v={v}\"")
    html = html.replace("app.js\"", f"app.js?v={v}\"")
    return HTMLResponse(html)


# ── 인증 (카카오 소셜 로그인) ─────────────────────────────────────────────────

@app.get("/auth/kakao/login", summary="카카오 로그인 페이지로 리다이렉트")
def kakao_login():
    if not KAKAO_CLIENT_ID:
        raise HTTPException(status_code=503, detail="카카오 앱 키가 설정되지 않았습니다. (KAKAO_CLIENT_ID 환경변수 필요)")
    return RedirectResponse(get_kakao_login_url())


@app.get("/auth/kakao/callback", summary="카카오 OAuth 콜백")
async def kakao_callback(code: str):
    access_token = await get_kakao_token(code)
    profile      = await get_kakao_profile(access_token)
    user_id      = upsert_user(
        kakao_id      = profile["kakao_id"],
        nickname      = profile["nickname"],
        profile_image = profile["profile_image"],
        email         = profile["email"],
    )
    token = create_access_token(user_id, profile["nickname"])
    return RedirectResponse(f"/?token={token}")


@app.get("/auth/me", summary="현재 로그인 사용자 정보")
def auth_me(current_user: dict = Depends(get_current_user)):
    user = get_user_by_id(current_user["user_id"])
    if not user:
        raise HTTPException(status_code=404, detail="사용자 없음")
    return user


@app.post("/auth/dev-login", summary="개발용 테스트 로그인 (카카오 앱 없이 테스트)")
def dev_login(nickname: str = "테스트사용자"):
    """카카오 앱 연동 전 로컬 개발/테스트용. 프로덕션에서는 비활성화 권장."""
    dev_kakao_id = f"dev_{nickname}"
    user_id = upsert_user(
        kakao_id      = dev_kakao_id,
        nickname      = nickname,
        profile_image = "",
        email         = "",
    )
    token = create_access_token(user_id, nickname)
    return {"access_token": token, "token_type": "bearer", "nickname": nickname}


# ── 내부 헬퍼: 소유권 검증 ─────────────────────────────────────────────────────

def _check_owner(taxpayer_id: int, user_id: int):
    conn = get_conn()
    tp = conn.execute("SELECT user_id FROM taxpayers WHERE id=?", (taxpayer_id,)).fetchone()
    conn.close()
    if not tp:
        raise HTTPException(status_code=404, detail="납세자 없음")
    if tp["user_id"] != user_id:
        raise HTTPException(status_code=403, detail="접근 권한이 없습니다.")


def _get_taxpayer_data(taxpayer_id: int) -> dict:
    """내부 전용: 인증 없이 납세자 데이터 조회 (공유 링크 등에서 사용)"""
    conn = get_conn()
    tp = conn.execute("SELECT * FROM taxpayers WHERE id=?", (taxpayer_id,)).fetchone()
    if not tp:
        conn.close()
        raise HTTPException(status_code=404, detail="납세자 없음")

    result = dict(tp)
    for table in [
        "businesses", "other_incomes", "deductions",
        "penalty_taxes", "tax_history", "income_rate_history",
        "sg_expenses", "credit_card_usage",
    ]:
        rows = conn.execute(
            f"SELECT * FROM {table} WHERE taxpayer_id=?", (taxpayer_id,)
        ).fetchall()
        result[table] = [dict(r) for r in rows]

    conn.close()
    return result


# ── PDF 업로드 & DB 저장 ───────────────────────────────────────────────────────

@app.post("/upload", summary="PDF 업로드 및 DB 저장")
async def upload_pdf(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="PDF 파일만 업로드 가능합니다.")

    safe_name = f"{uuid.uuid4().hex}_{file.filename}"
    save_path = os.path.join(UPLOAD_DIR, safe_name)
    with open(save_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    try:
        data = parse_tax_pdf(save_path)
    except Exception as e:
        os.remove(save_path)
        raise HTTPException(status_code=422, detail=f"PDF 파싱 실패: {e}")

    conn = get_conn()
    try:
        cur = conn.cursor()

        tp = data["taxpayer"]
        cur.execute("""
            INSERT INTO taxpayers
              (tax_year, name, birth_date, guide_type, bookkeeping_obligation,
               estimated_expense_rate, payment_extension, ars_auth_number,
               religion_income, pdf_filename, user_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """, (
            tp.get("tax_year"), tp.get("name"), tp.get("birth_date"),
            tp.get("guide_type"), tp.get("bookkeeping_obligation"),
            tp.get("estimated_expense_rate"), tp.get("payment_extension"),
            tp.get("ars_auth_number"), tp.get("religion_income", "X"),
            file.filename,
            current_user["user_id"],
        ))
        taxpayer_id = cur.lastrowid

        for b in data["businesses"]:
            cur.execute("""
                INSERT INTO businesses
                  (taxpayer_id, business_reg_no, business_name, income_type_code,
                   industry_code, business_type, bookkeeping_obligation,
                   expense_rate_type, revenue,
                   std_expense_rate_general, std_expense_rate_own,
                   simple_expense_rate_general, simple_expense_rate_own)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                taxpayer_id,
                b.get("business_reg_no"), b.get("business_name"),
                b.get("income_type_code"), b.get("industry_code"),
                b.get("business_type"), b.get("bookkeeping_obligation"),
                b.get("expense_rate_type"), b.get("revenue"),
                b.get("std_expense_rate_general"), b.get("std_expense_rate_own"),
                b.get("simple_expense_rate_general"), b.get("simple_expense_rate_own"),
            ))

        for oi in data["other_incomes"]:
            cur.execute(
                "INSERT INTO other_incomes (taxpayer_id, income_type, has_data) VALUES (?,?,?)",
                (taxpayer_id, oi["income_type"], oi["has_data"])
            )

        for d in data["deductions"]:
            cur.execute(
                "INSERT INTO deductions (taxpayer_id, category, item_name, amount) VALUES (?,?,?,?)",
                (taxpayer_id, d["category"], d["item_name"], d.get("amount"))
            )

        for p in data["penalty_taxes"]:
            cur.execute("""
                INSERT INTO penalty_taxes (taxpayer_id, penalty_type, detail_type, count, amount)
                VALUES (?,?,?,?,?)
            """, (taxpayer_id, p["penalty_type"], p.get("detail_type"), p.get("count"), p.get("amount")))

        for h in data["tax_history"]:
            cur.execute("""
                INSERT INTO tax_history
                  (taxpayer_id, attribution_year, total_income, income_deduction,
                   taxable_income, tax_rate, calculated_tax, deduction_tax,
                   determined_tax, effective_tax_rate)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (
                taxpayer_id,
                h.get("attribution_year"), h.get("total_income"), h.get("income_deduction"),
                h.get("taxable_income"), h.get("tax_rate"), h.get("calculated_tax"),
                h.get("deduction_tax"), h.get("determined_tax"), h.get("effective_tax_rate"),
            ))

        for ir in data["income_rate_history"]:
            cur.execute("""
                INSERT INTO income_rate_history
                  (taxpayer_id, business_reg_no, business_name,
                   attribution_year, revenue, necessary_expenses, income, income_rate)
                VALUES (?,?,?,?,?,?,?,?)
            """, (
                taxpayer_id,
                ir.get("business_reg_no"), ir.get("business_name"),
                ir.get("attribution_year"), ir.get("revenue"),
                ir.get("necessary_expenses"), ir.get("income"), ir.get("income_rate"),
            ))

        for sg in data["sg_expenses"]:
            cur.execute("""
                INSERT INTO sg_expenses
                  (taxpayer_id, analysis_year, account_code, account_name,
                   amount, company_rate, industry_avg_rate)
                VALUES (?,?,?,?,?,?,?)
            """, (
                taxpayer_id,
                sg.get("analysis_year"), sg.get("account_code"), sg.get("account_name"),
                sg.get("amount"), sg.get("company_rate"), sg.get("industry_avg_rate"),
            ))

        for cc in data["credit_card_usage"]:
            cur.execute("""
                INSERT INTO credit_card_usage (taxpayer_id, usage_year, category, count, amount)
                VALUES (?,?,?,?,?)
            """, (taxpayer_id, cc.get("usage_year"), cc["category"], cc.get("count"), cc.get("amount")))

        conn.commit()
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"DB 저장 실패: {e}")
    finally:
        conn.close()

    return {
        "status": "success",
        "taxpayer_id": taxpayer_id,
        "parsed": {
            "businesses":          len(data["businesses"]),
            "other_incomes":       len(data["other_incomes"]),
            "deductions":          len(data["deductions"]),
            "penalty_taxes":       len(data["penalty_taxes"]),
            "tax_history":         len(data["tax_history"]),
            "income_rate_history": len(data["income_rate_history"]),
            "sg_expenses":         len(data["sg_expenses"]),
            "credit_card_usage":   len(data["credit_card_usage"]),
        },
        "data": data,
    }


# ── 조회 API ──────────────────────────────────────────────────────────────────

@app.get("/taxpayers", summary="납세자 목록 조회 (본인 데이터만)")
def list_taxpayers(current_user: dict = Depends(get_current_user)):
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM taxpayers WHERE user_id=? ORDER BY id DESC",
        (current_user["user_id"],)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.get("/taxpayers/{taxpayer_id}", summary="납세자 상세 조회")
def get_taxpayer(taxpayer_id: int, current_user: dict = Depends(get_current_user)):
    _check_owner(taxpayer_id, current_user["user_id"])
    return _get_taxpayer_data(taxpayer_id)


# ── 세금 자동 계산 ────────────────────────────────────────────────────────────

@app.get("/taxpayers/{taxpayer_id}/calculate", summary="종합소득세 자동 계산")
def calc_tax(taxpayer_id: int, current_user: dict = Depends(get_current_user)):
    _check_owner(taxpayer_id, current_user["user_id"])
    result = calculate_tax(taxpayer_id)
    if "error" in result:
        raise HTTPException(status_code=404, detail=result["error"])
    return result


# ── AI 분석 코멘트 ────────────────────────────────────────────────────────────

@app.get("/taxpayers/{taxpayer_id}/ai-analysis", summary="AI 분석 코멘트")
def ai_analysis(taxpayer_id: int, current_user: dict = Depends(get_current_user)):
    _check_owner(taxpayer_id, current_user["user_id"])
    result = generate_ai_analysis(taxpayer_id)
    if "error" in result:
        raise HTTPException(status_code=404, detail=result["error"])
    return result


# ── 공유 링크 ─────────────────────────────────────────────────────────────────

@app.post("/taxpayers/{taxpayer_id}/share", summary="공유 링크 토큰 생성")
def create_share(taxpayer_id: int, current_user: dict = Depends(get_current_user)):
    _check_owner(taxpayer_id, current_user["user_id"])

    token      = uuid.uuid4().hex[:16]
    expires_at = (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%dT%H:%M:%S")

    conn = get_conn()
    conn.execute(
        "INSERT OR REPLACE INTO share_tokens (token, taxpayer_id, expires_at) VALUES (?,?,?)",
        (token, taxpayer_id, expires_at),
    )
    conn.commit()
    conn.close()
    return {"token": token, "expires_at": expires_at}


@app.get("/share/{token}", summary="공유 링크로 데이터 조회 (인증 불필요)")
def get_share_data(token: str):
    conn = get_conn()
    row = conn.execute(
        "SELECT taxpayer_id, expires_at FROM share_tokens WHERE token=?", (token,)
    ).fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="유효하지 않은 공유 링크입니다.")
    if datetime.now().strftime("%Y-%m-%dT%H:%M:%S") > row["expires_at"]:
        raise HTTPException(status_code=410, detail="만료된 공유 링크입니다 (7일 초과).")
    return _get_taxpayer_data(row["taxpayer_id"])


# ── 엑셀 내보내기 ─────────────────────────────────────────────────────────────

def _xl_header(ws, headers: list, fill_color="1E40AF"):
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _xl_autowidth(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


@app.get("/taxpayers/{taxpayer_id}/export/excel", summary="엑셀 내보내기")
def export_excel(taxpayer_id: int, current_user: dict = Depends(get_current_user)):
    _check_owner(taxpayer_id, current_user["user_id"])

    conn = get_conn()
    tp = conn.execute("SELECT * FROM taxpayers WHERE id=?", (taxpayer_id,)).fetchone()
    if not tp:
        conn.close()
        raise HTTPException(status_code=404, detail="납세자 없음")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("01_기본정보")
    _xl_header(ws1, ["항목", "값"])
    labels = [("귀속연도", tp["tax_year"]), ("성명", tp["name"]),
              ("생년월일", tp["birth_date"]), ("안내유형", tp["guide_type"]),
              ("기장의무", tp["bookkeeping_obligation"]),
              ("추계시 적용경비율", tp["estimated_expense_rate"]),
              ("종교인기타 소득유무", tp["religion_income"]),
              ("업로드일시", tp["uploaded_at"]), ("원본 파일명", tp["pdf_filename"])]
    for i, (k, v) in enumerate(labels, 2):
        ws1.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws1.cell(row=i, column=2, value=v)
    _xl_autowidth(ws1)

    ws2 = wb.create_sheet("02_사업장수입")
    biz_headers = ["사업자번호", "상호", "수입종류", "업종코드", "사업형태",
                   "기장의무", "경비율", "수입금액(원)", "기준경비율일반(%)", "단순경비율일반(%)"]
    _xl_header(ws2, biz_headers)
    for r_idx, row in enumerate(conn.execute(
        "SELECT * FROM businesses WHERE taxpayer_id=?", (taxpayer_id,)
    ).fetchall(), 2):
        vals = [row["business_reg_no"], row["business_name"], row["income_type_code"],
                row["industry_code"], row["business_type"], row["bookkeeping_obligation"],
                row["expense_rate_type"], row["revenue"],
                row["std_expense_rate_general"], row["simple_expense_rate_general"]]
        for c_idx, v in enumerate(vals, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=v)
            if c_idx == 8:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
    _xl_autowidth(ws2)

    ws3 = wb.create_sheet("03_종합소득세_3년")
    hist_rows = conn.execute(
        "SELECT * FROM tax_history WHERE taxpayer_id=? ORDER BY attribution_year", (taxpayer_id,)
    ).fetchall()
    years = [str(r["attribution_year"]) + "귀속" for r in hist_rows]
    _xl_header(ws3, ["구분(천원)"] + years)
    fields = [("종합소득금액", "total_income"), ("소득공제", "income_deduction"),
              ("과세표준", "taxable_income"), ("세율(%)", "tax_rate"),
              ("산출세액", "calculated_tax"), ("공제·감면세액", "deduction_tax"),
              ("결정세액", "determined_tax"), ("실효세율(%)", "effective_tax_rate")]
    for r_i, (label, key) in enumerate(fields, 2):
        ws3.cell(row=r_i, column=1, value=label).font = Font(bold=True)
        for c_i, row in enumerate(hist_rows, 2):
            cell = ws3.cell(row=r_i, column=c_i, value=row[key])
            cell.alignment = Alignment(horizontal="right")
            if "율" in label:
                cell.number_format = '0.00"%"'
            else:
                cell.number_format = '#,##0'
    _xl_autowidth(ws3)

    ws4 = wb.create_sheet("04_신고소득률")
    _xl_header(ws4, ["귀속연도", "상호", "수입금액(천원)", "필요경비(천원)", "소득금액(천원)", "소득률(%)"])
    for r_i, row in enumerate(conn.execute(
        "SELECT * FROM income_rate_history WHERE taxpayer_id=? ORDER BY attribution_year", (taxpayer_id,)
    ).fetchall(), 2):
        ws4.cell(row=r_i, column=1, value=row["attribution_year"])
        ws4.cell(row=r_i, column=2, value=row["business_name"])
        for c_i, key in enumerate(["revenue", "necessary_expenses", "income"], 3):
            ws4.cell(row=r_i, column=c_i, value=row[key]).number_format = '#,##0'
        ws4.cell(row=r_i, column=6, value=row["income_rate"]).number_format = '0.00"%"'
    _xl_autowidth(ws4)

    ws5 = wb.create_sheet("05_판관비분석")
    _xl_header(ws5, ["계정과목코드", "계정과목명", "금액(천원)", "당해업체(%)", "업종평균(%)"])
    for r_i, row in enumerate(conn.execute(
        "SELECT * FROM sg_expenses WHERE taxpayer_id=?", (taxpayer_id,)
    ).fetchall(), 2):
        ws5.cell(row=r_i, column=1, value=row["account_code"])
        ws5.cell(row=r_i, column=2, value=row["account_name"])
        ws5.cell(row=r_i, column=3, value=row["amount"]).number_format = '#,##0'
        ws5.cell(row=r_i, column=4, value=row["company_rate"]).number_format = '0.00"%"'
        ws5.cell(row=r_i, column=5, value=row["industry_avg_rate"]).number_format = '0.00"%"'
    _xl_autowidth(ws5)

    ws6 = wb.create_sheet("06_공제내역")
    _xl_header(ws6, ["구분", "항목명", "금액(원)"])
    for r_i, row in enumerate(conn.execute(
        "SELECT * FROM deductions WHERE taxpayer_id=?", (taxpayer_id,)
    ).fetchall(), 2):
        ws6.cell(row=r_i, column=1, value=row["category"]).font = Font(bold=True)
        ws6.cell(row=r_i, column=2, value=row["item_name"])
        ws6.cell(row=r_i, column=3, value=row["amount"]).number_format = '#,##0'
    _xl_autowidth(ws6)

    ws7 = wb.create_sheet("07_신용카드")
    _xl_header(ws7, ["구분", "건수", "금액(원)"])
    for r_i, row in enumerate(conn.execute(
        "SELECT * FROM credit_card_usage WHERE taxpayer_id=?", (taxpayer_id,)
    ).fetchall(), 2):
        ws7.cell(row=r_i, column=1, value=row["category"])
        ws7.cell(row=r_i, column=2, value=row["count"])
        ws7.cell(row=r_i, column=3, value=row["amount"]).number_format = '#,##0'
    _xl_autowidth(ws7)

    ws8 = wb.create_sheet("08_가산세")
    _xl_header(ws8, ["가산세 항목", "세부 구분", "건수", "금액(원)"])
    for r_i, row in enumerate(conn.execute(
        "SELECT * FROM penalty_taxes WHERE taxpayer_id=?", (taxpayer_id,)
    ).fetchall(), 2):
        ws8.cell(row=r_i, column=1, value=row["penalty_type"])
        ws8.cell(row=r_i, column=2, value=row["detail_type"])
        ws8.cell(row=r_i, column=3, value=row["count"])
        amt_cell = ws8.cell(row=r_i, column=4, value=row["amount"])
        if row["amount"] is not None:
            amt_cell.number_format = '#,##0'
    _xl_autowidth(ws8)

    conn.close()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    name     = tp["name"] or "납세자"
    year     = tp["tax_year"] or ""
    filename = f"ByeTax_{name}_{year}귀속.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename, safe='')}"},
    )
